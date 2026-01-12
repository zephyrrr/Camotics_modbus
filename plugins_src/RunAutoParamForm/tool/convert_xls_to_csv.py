import openpyxl
import os
import csv

def convert(file_path):
    root_path = os.path.dirname(file_path)
    #root_path = 'C:\\OneDrive\\doing\\cnc\\Camotics_modbus\\plugins\\RunAutoParamForm'
    #file_path = os.path.join(root_path, '自动参数模型7.xlsx')
    #os.chdir(root_path)

    if not os.path.exists(file_path):
        print(f"Error: The file '{file_path}' was not found.")

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    sheet_names = '''通用
深腔
大面积
骨位
插件
多脚
螺牙
蜗杆
清角
进胶口'''.split('\n')
    #sheet_names = ('通用', )
    output_csv_path = os.path.join(root_path, '库.csv')
    with open(output_csv_path, 'w', newline='', encoding='utf-8-sig') as csv_file:
        csv_writer = csv.writer(csv_file)

        for sheet_name in sheet_names:
            if not sheet_name:
                continue
            # Select the first worksheet by its name
            # You can also use workbook.active to get the active sheet
            sheet = workbook[sheet_name]

            # Read data from a specific cell
            #cell_a1_value = sheet['B'].value
            #print(f"Value in cell A1: {cell_a1_value}")
            # # Read data from a range of cells
            # print("\nReading data from a range:")
            # for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=34):
            #     for cell in row:
            #         print(f"{cell.coordinate}: {cell.value}", end="\t")
            #     print()

            column_cnt = 34
            empty_row_cnt = 0
            for row_index, row in enumerate(sheet.iter_rows(), start=1):
                if row_index == 1:
                    if sheet_name == sheet_names[0]:
                        row_data = [cell.value if cell.value is not None else '' for cell in row[:column_cnt]]
                        csv_writer.writerow(row_data)
                    continue
                
                row_data = []
                for cell_index, cell in enumerate(row[:column_cnt]):
                    if cell_index <= 20 and cell.data_type == 'n' and cell.value:
                        row_data.append(round(cell.value))
                    else:
                        row_data.append(cell.value)
                if all(value is None or value == '' for value in row_data):
                    empty_row_cnt += 1
                    if empty_row_cnt >= 10:
                        break
                if row_data[1] is None or not str(row_data[1]).strip():
                    continue
                empty_row_cnt = 0
                # if row_index == 4:
                #     break
                #print(row_data)
                # Write the row to the CSV file
                csv_writer.writerow(row_data)

    workbook.close()
    print(f"Successfully converted.")


def has_duplicate_rows(file_path = '库.csv'):
    """
    Checks if a CSV file has any duplicate rows.

    Args:
        file_path (str): The path to the CSV file.

    Returns:
        bool: True if a duplicate row is found, False otherwise.
    """
    try:
        seen_rows = set()
        with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            for row_index, row in enumerate(reader):
                # Convert the list of cell values to a tuple so it can be added to a set
                row_tuple = tuple(row)
                if row_tuple in seen_rows:
                    print(f"Duplicate row found: {','.join(row)}, {row_index}")
                    #return True
                    continue
                seen_rows.add(row_tuple)
        return False
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
        return False
    except Exception as e:
        print(f"An error occurred: {e}")
        return False
		
if __name__ == "__main__":
	convert('./自动参数模型7.xlsx')